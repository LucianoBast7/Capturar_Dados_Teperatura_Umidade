import requests
from openpyxl import Workbook, load_workbook
import datetime
import tkinter as tk
from tkinter import messagebox

# Configuração da API
api_key = "c5380ad59a64e1300b251335eac748fa"
cidade = "São Paulo"
url = f"http://api.openweathermap.org/data/2.5/weather?q={cidade}&appid={api_key}&units=metric"

# Função para capturar e salvar os dados de temperatura e umidade
def captura_dados():
    try:
        # Faz a requisição para a API
        response = requests.get(url)
        dados = response.json()

        # Extrai temperatura e umidade
        temperatura = dados['main']['temp']
        umidade = dados['main']['humidity']
        data_hora = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Define o status da umidade
        status_umidade = "Alta" if umidade > 70 else "Normal" if umidade > 30 else "Baixa"

        # Verifica se o arquivo Excel já existe
        try:
            workbook = load_workbook("historico_clima.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Data/Hora", "Temperatura (°C)", "Status da Umidade"])

        # Adiciona os dados na planilha
        sheet.append([data_hora, temperatura, status_umidade])
        workbook.save("historico_clima.xlsx")

        # Confirmação ao usuário
        messagebox.showinfo("Sucesso", f"Dados salvos: {data_hora}\nTemperatura: {temperatura}°C\nUmidade: {status_umidade}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao capturar dados: {e}")

# Interface do usuário com Tkinter
root = tk.Tk()
root.title("Captura de Dados do Clima")
root.geometry("400x200")

# Botão para iniciar a captação
botao_capturar = tk.Button(root, text="Capturar Dados de Clima", command=captura_dados, padx=20, pady=10)
botao_capturar.pack(pady=20)

root.mainloop()

